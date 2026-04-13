#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl


def round2(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value) + 1e-12, 2)
    except Exception:
        return None


def normalize_text(value: Any) -> str:
    if value is None:
        return ''
    text = str(value).replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def excel_value(ws, cell_ref: str):
    return ws[cell_ref].value


def fmt_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    text = normalize_text(value)
    if not text:
        return ''
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d.%m.%Y', '%d.%m.%Y г.'):
        try:
            return datetime.strptime(text, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return text


def formula_or_value(ws_formula, ws_values, cell_ref: str):
    raw_formula = ws_formula[cell_ref].value
    raw_value = ws_values[cell_ref].value
    return raw_value if raw_value is not None else raw_formula


def infer_category(name: str, note: str) -> str:
    source = f'{name} {note}'.lower()
    if re.search(r'арматур|металл|сталь', source):
        return 'metal'
    if 'каркас' in source:
        return 'frame'
    if 'бетон' in source:
        return 'concrete'
    if re.search(r'материал|щебень|песок|гидрошпон|плита|пвх|gener', source):
        return 'material'
    if re.search(r'проч|пр.мат', source):
        return 'misc'
    return 'work'


def parse_common(ws) -> dict[str, Any]:
    return {
        'okudKs2': normalize_text(excel_value(ws, 'J5')),
        'okudKs3': '0322001',
        'developerName': normalize_text(excel_value(ws, 'D7')),
        'developerOkpo': normalize_text(excel_value(ws, 'J7')),
        'techCustomerName': normalize_text(excel_value(ws, 'D9')),
        'techCustomerOkpo': normalize_text(excel_value(ws, 'J9')),
        'contractorName': normalize_text(excel_value(ws, 'D11')),
        'contractorOkpo': normalize_text(excel_value(ws, 'J11')),
        'constructionObject': normalize_text(excel_value(ws, 'D13')),
        'objectName': normalize_text(excel_value(ws, 'D13')),
        'contractNumber': normalize_text(excel_value(ws, 'J17')),
        'contractDate': normalize_text(excel_value(ws, 'J18')),
        'operationType': normalize_text(excel_value(ws, 'J19')),
        'currencyCode': '643',
        'currencyName': 'Российский рубль',
        'showDocumentHeaders': True,
        'showDocumentSignatures': True,
    }


def parse_signers_from_sheet(ws, common: dict[str, Any]) -> None:
    for row in range(1, ws.max_row + 1):
        marker = normalize_text(ws.cell(row, 2).value)
        position = normalize_text(ws.cell(row, 4).value)
        name = normalize_text(ws.cell(row, 8).value)
        if marker == 'Сдал:':
            common['contractorSignerPosition'] = position
            common['contractorSignerName'] = name
        elif marker == 'Принял:':
            common['customerSignerPosition'] = position
            common['customerSignerName'] = name
        elif marker == 'Проверил:':
            common['techCustomerSignerPosition'] = position
            common['techCustomerSignerName'] = name


def parse_ks2_sheet(ws_formula, ws_values) -> dict[str, Any]:
    title = ws_formula.title
    sheet = {
        'id': re.sub(r'[^a-zA-Z0-9а-яА-Я_-]+', '-', title).strip('-').lower(),
        'title': title,
        'documentNumber': normalize_text(excel_value(ws_formula, 'G23')),
        'documentDate': fmt_date(excel_value(ws_formula, 'H23')),
        'periodFrom': fmt_date(excel_value(ws_formula, 'I23')),
        'periodTo': fmt_date(excel_value(ws_values, 'J23') or excel_value(ws_formula, 'H23')),
        'basis': normalize_text(excel_value(ws_formula, 'C25')),
        'vatRate': 22 if '(22%)' in normalize_text(excel_value(ws_formula, 'G27')) else 20,
        'rows': [],
    }

    totals_row = None
    for row in range(29, ws_formula.max_row + 1):
        name = normalize_text(ws_formula.cell(row, 4).value)
        if name == 'ВСЕГО по Акту:':
            totals_row = row
            break

        code = normalize_text(ws_formula.cell(row, 1).value)
        line_no = normalize_text(ws_formula.cell(row, 2).value)
        estimate_no = normalize_text(ws_formula.cell(row, 3).value)
        unit = normalize_text(ws_formula.cell(row, 5).value)
        quantity = formula_or_value(ws_formula, ws_values, f'F{row}')
        price = formula_or_value(ws_formula, ws_values, f'G{row}')
        amount = formula_or_value(ws_formula, ws_values, f'H{row}')
        unit_consumption = formula_or_value(ws_formula, ws_values, f'I{row}')
        note = normalize_text(ws_formula.cell(row, 10).value)

        if name and not any([code, line_no, estimate_no, unit, quantity, price, amount]):
            row_type = 'note' if name.lower().startswith('корректировка') else 'section'
            sheet['rows'].append({
                'type': row_type,
                'code': code,
                'lineNo': line_no,
                'estimateNo': estimate_no,
                'name': name,
                'unit': unit,
                'quantity': None,
                'price': None,
                'amount': None,
                'unitConsumption': None,
                'note': note,
                'category': 'misc',
            })
            continue

        if not name:
            continue

        category = infer_category(name, note)
        sheet['rows'].append({
            'type': 'item',
            'code': code,
            'lineNo': line_no,
            'estimateNo': estimate_no,
            'name': name,
            'unit': unit,
            'quantity': round2(quantity),
            'price': round2(price),
            'amount': round2(amount),
            'unitConsumption': round2(unit_consumption),
            'note': note,
            'category': category,
        })

    gross = round2(ws_values[f'H{totals_row}'].value if totals_row else None) or 0.0
    vat = round2(ws_values[f'H{totals_row + 1}'].value if totals_row else None) or 0.0
    breakdown = []
    for r in range((totals_row or 0), min((totals_row or 0) + 8, ws_formula.max_row + 1)):
        label = normalize_text(ws_formula.cell(r, 12).value)
        amount = round2(ws_values.cell(r, 11).value)
        if label:
            breakdown.append({'label': label, 'amount': amount or 0.0})

    sheet['totals'] = {'gross': gross, 'vat': vat, 'breakdown': breakdown}
    return sheet


def parse_holdbacks(ws_formula, ws_values) -> dict[str, Any]:
    rows = []
    for row in range(7, 24):
        name = normalize_text(ws_formula.cell(row, 1).value)
        if not name and not normalize_text(ws_formula.cell(row, 4).value):
            continue
        rows.append({
            'name': name,
            'ks2Amount': round2(ws_values.cell(row, 2).value),
            'materialsUsed': round2(ws_values.cell(row, 3).value),
            'advanceReceived': round2(ws_values.cell(row, 4).value),
            'advanceDoc': normalize_text(ws_formula.cell(row, 5).value),
            'previousBalance': round2(ws_values.cell(row, 6).value),
            'closingAmount': round2(ws_values.cell(row, 7).value),
            'nextBalance': round2(ws_values.cell(row, 8).value),
            'retentionAmount': round2(ws_values.cell(row, 9).value),
            'payableAmount': round2(ws_values.cell(row, 10).value),
            'comment': '',
        })

    totals = {
        'ks2Amount': round2(ws_values['B24'].value),
        'materialsUsed': round2(ws_values['C24'].value),
        'advanceReceived': round2(ws_values['D24'].value),
        'previousBalance': round2(ws_values['F24'].value),
        'closingAmount': round2(ws_values['G24'].value),
        'nextBalance': round2(ws_values['H24'].value),
        'retentionAmount': round2(ws_values['I24'].value),
        'payableAmount': round2(ws_values['J24'].value),
    }
    return {'title': 'Удержания', 'rows': rows, 'totals': totals}


def build_state(xlsx_path: Path) -> dict[str, Any]:
    wb_formula = openpyxl.load_workbook(xlsx_path, data_only=False)
    wb_values = openpyxl.load_workbook(xlsx_path, data_only=True)

    first_ks2 = next(ws for ws in wb_formula.worksheets if ws.title.startswith('КС-2'))
    common = parse_common(first_ks2)

    ks2_sheets = []
    for ws_formula, ws_values in zip(wb_formula.worksheets, wb_values.worksheets):
        if ws_formula.title.startswith('КС-2'):
            parse_signers_from_sheet(ws_formula, common)
            ks2_sheets.append(parse_ks2_sheet(ws_formula, ws_values))

    hold_ws_formula = wb_formula['Удержания'] if 'Удержания' in wb_formula.sheetnames else None
    hold_ws_values = wb_values['Удержания'] if 'Удержания' in wb_values.sheetnames else None
    holdbacks = parse_holdbacks(hold_ws_formula, hold_ws_values) if hold_ws_formula and hold_ws_values else {'title': 'Удержания', 'rows': [], 'totals': {}}

    xml_manual = {
        'economicSubjectName': common.get('contractorName', ''),
        'estimateVersionCode': '1',
        'supplementDocType': '',
        'supplementDocNumber': '',
        'supplementDocDate': '',
        'contractorInn': '',
        'customerInn': '',
        'developerPostalIndex': '',
        'developerRegionCode': '77',
        'contractorPostalIndex': '',
        'contractorRegionCode': '77',
        'correctionNumber': '',
        'correctionDate': '',
        'signerName': common.get('contractorSignerName', ''),
        'signerPosition': common.get('contractorSignerPosition', ''),
        'signerStatus': '1',
        'signatureType': '1',
    }

    return {
        'meta': {
            'sourceWorkbook': xlsx_path.name,
            'description': f'Импортировано из {xlsx_path.name}',
        },
        'common': common,
        'ks2Sheets': ks2_sheets,
        'ks3': {
            'rows': [],
            'totals': {},
            'documentNumber': ks2_sheets[0]['documentNumber'] if ks2_sheets else '',
            'documentDate': ks2_sheets[0]['documentDate'] if ks2_sheets else '',
            'periodFrom': ks2_sheets[0]['periodFrom'] if ks2_sheets else '',
            'periodTo': ks2_sheets[0]['periodTo'] if ks2_sheets else '',
        },
        'holdbacks': holdbacks,
        'xmlExtras': {
            'generated': {
                'knd': '1110335',
                'formatVersion': '1.00',
                'programVersion': 'prototype-0.1.0',
                'fileDate': datetime.now().strftime('%Y-%m-%d'),
                'fileTime': datetime.now().strftime('%H:%M:%S'),
            },
            'constants': {
                'isGovMunicipal': '0',
                'vatCalcInTotalOnly': '1',
                'cumulativeMode': '1',
                'priceIndexYear': '0000',
                'requiresSettlementApproval': '1',
            },
            'manual': xml_manual,
            'traceableGoods': [],
        },
        'ui': {
            'activePane': 'requisites',
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument('xlsx_path')
    parser.add_argument('-o', '--output', required=True)
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx_path).expanduser().resolve()
    output = Path(args.output).expanduser().resolve()
    state = build_state(xlsx_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding='utf-8')
    print(output)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
