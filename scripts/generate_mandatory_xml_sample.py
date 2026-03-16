#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from app.xsd_utils import FieldDef, build_xml_from_values, parse_xsd_fields, validate_xml
XSD_PATH = ROOT / "nalog docs" / "ON_AKTREZRABP_1_971_01_01_00_03.xsd"
MANDATORY_MD = ROOT / "nalog docs" / "обязательные поля 1110335 (безусловные).md"
SAMPLE_XLSX = ROOT / "nalog docs" / "пример заполнения.xlsx"
OUT_XML = ROOT / "output" / "sample_mandatory_1110335.xml"
OUT_REPORT = ROOT / "nalog docs" / "пример xml 1110335 - отчет xsd.md"
OUT_VALUES_JSON = ROOT / "output" / "sample_mandatory_1110335.values.json"


def parse_md_table(path: Path, min_cols: int) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.startswith("|"):
            continue
        parts = [p.strip().replace("\\|", "|") for p in re.split(r"(?<!\\)\|", line)[1:-1]]
        if len(parts) < min_cols:
            continue
        if parts[0] in {"Таблица", "№", "---"}:
            continue
        if all(p.startswith("-") for p in parts):
            continue
        rows.append(parts)
    return rows


def fmt_date(v) -> str:
    if hasattr(v, "strftime"):
        return v.strftime("%d.%m.%Y")

    s = str(v).strip()
    s = re.sub(r"\s*г\.?$", "", s).strip()

    # yyyy-mm-dd -> dd.mm.yyyy
    m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(3)}.{m.group(2)}.{m.group(1)}"

    # dd.mm.yyyy accepted as-is
    m = re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", s)
    if m:
        return s

    return s


def fmt_amount(v) -> str:
    try:
        num = float(str(v).replace(" ", "").replace(",", "."))
        return f"{num:.2f}"
    except Exception:
        return str(v).strip()


def parse_excel_seed(path: Path) -> dict[str, str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    seed: dict[str, str] = {}

    # Core header values from sample sheet
    seed["doc_num"] = str(ws.cell(23, 7).value or "1").strip()
    seed["doc_date"] = fmt_date(ws.cell(23, 8).value or datetime.now())
    seed["period_from"] = fmt_date(ws.cell(23, 9).value or datetime.now())
    seed["period_to"] = fmt_date(ws.cell(23, 10).value or datetime.now())

    seed["builder_name"] = str(ws.cell(7, 4).value or "ООО \"Пример\"").strip()
    seed["customer_name"] = str(ws.cell(9, 4).value or "ООО \"Заказчик\"").strip()
    seed["contractor_name"] = str(ws.cell(11, 4).value or "ООО \"Подрядчик\"").strip()

    seed["site_name"] = str(ws.cell(13, 4).value or "Объект строительства").strip()
    seed["object_name"] = str(ws.cell(15, 4).value or seed["site_name"]).strip()

    seed["contract_no"] = str(ws.cell(17, 10).value or "ДОГ-001").strip()
    seed["contract_date"] = fmt_date(ws.cell(18, 10).value or datetime.now())

    seed["operation_text"] = str(ws.cell(24, 4).value or "О ПРИЕМКЕ ВЫПОЛНЕННЫХ РАБОТ").strip()
    seed["basis_text"] = str(ws.cell(25, 3).value or "Основание").strip()

    seed["section_name"] = str(ws.cell(29, 4).value or "Раздел 1").strip()
    seed["work_name"] = str(ws.cell(30, 4).value or "Выполнение работ").strip()

    seed["price"] = fmt_amount(ws.cell(30, 7).value or "100000.00")
    seed["line_total"] = fmt_amount(ws.cell(30, 8).value or "100000.00")
    seed["total"] = fmt_amount(ws.cell(39, 8).value or "100000.00")
    seed["vat_sum"] = fmt_amount(ws.cell(40, 8).value or "20000.00")

    # Signers from sample rows (fallback to generated names if empty)
    seed["sign_contractor"] = str(ws.cell(43, 8).value or "Иванов И.И.").strip()
    seed["sign_customer"] = str(ws.cell(46, 8).value or "Петров П.П.").strip()

    return seed


def split_name(full_name: str) -> tuple[str, str, str]:
    text = re.sub(r"\s+", " ", (full_name or "").replace(".", " ").strip())
    parts = [p for p in text.split(" ") if p]
    if not parts:
        return "Иванов", "Иван", ""
    if len(parts) == 1:
        return parts[0], "Иван", ""
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return parts[0], parts[1], " ".join(parts[2:])


def build_path_mapping(mandatory_rows: list[list[str]], fields: list[FieldDef]) -> tuple[list[dict], list[dict]]:
    mapped: list[dict] = []
    unmapped: list[dict] = []

    for row in mandatory_rows:
        tab, block, name, code, req, note = row[:6]
        code = code.strip()
        if not code:
            unmapped.append({"tab": tab, "block": block, "name": name, "code": code, "reason": "no code"})
            continue

        cands = [f.path for f in fields if f.path.endswith(f"/@{code}") or f.path.endswith(f"/{code}")]
        if not cands:
            unmapped.append({"tab": tab, "block": block, "name": name, "code": code, "reason": "no candidates"})
            continue

        block_match = re.search(r"\(([^()]+)\)\s*$", block)
        block_code = block_match.group(1) if block_match else None
        if block_code:
            pri = [p for p in cands if f"/{block_code}/" in p or p.endswith(f"/{block_code}") or p.endswith(f"/@{block_code}")]
            if pri:
                cands = pri

        # Prefer contractor document branch + shorter path
        cands = sorted(cands, key=lambda p: (0 if "/Файл/Документ/" in p else 1, len(p), p))
        chosen = cands[0]

        mapped.append(
            {
                "tab": tab,
                "block": block,
                "name": name,
                "code": code,
                "req": req,
                "note": note,
                "path": chosen,
                "candidate_count": len(cands),
            }
        )

    return mapped, unmapped


def choose_value(path: str, code: str, name: str, note: str, field: FieldDef | None, seed: dict[str, str], now: datetime) -> tuple[str, str]:
    # source labels: excel | generated | reference | constant | manual

    # 1) Explicit constants
    constants = {
        "/Файл/@ВерсФорм": "1.00",
        "/Файл/Документ/@КНД": "1110335",
        "/Файл/Документ/СвАктСдПр/@КодОКВДог": "643",
        "/Файл/Документ/СвАктСдПр/ДенИзм/@КодОКВ": "643",
        "/Файл/Документ/СвАктСдПр/ОсновСтроит/@ПрГосМун": "0",
        "/Файл/Документ/НастрФормДок/@ПрНДСВИтог": "0",
        "/Файл/Документ/НастрФормДок/@ПрНакИтог": "0",
        "/Файл/Документ/НастрФормДок/@ПрИндЦен": "0000",
        "/Файл/Документ/НастрФормДок/@ПрСведРасчСогл": "0",
    }
    if path in constants:
        return constants[path], "constant"

    # 2) Generated
    if path == "/Файл/@ИдФайл":
        file_id = f"ON_AKTREZRABP_0000000000000_0000000000000_{now.strftime('%Y%m%d')}_{now.strftime('%H%M%S%f')}"
        return file_id, "generated"
    if path == "/Файл/@ВерсПрог":
        return "KC2-KC3-XML-webapp", "generated"
    if path == "/Файл/Документ/@ДатаИнфПодр":
        return now.strftime("%d.%m.%Y"), "generated"
    if path == "/Файл/Документ/@ВремИнфПодр":
        return now.strftime("%H:%M:%S"), "generated"
    if path.endswith("@НаимЕдИзмПрослеж"):
        return "шт", "generated"
    if path.endswith("/НетИзмКол"):
        return "без изм", "generated"

    # 3) Excel-backed fields
    if path.endswith("@НомерДок") and "/СвАктСдПр/" in path:
        return seed["doc_num"], "excel"
    if path.endswith("@ДатаДок") and "/СвАктСдПр/" in path:
        return seed["doc_date"], "excel"
    if path.endswith("@НаимОб"):
        return seed["object_name"], "excel"
    if path.endswith("@НаимЭкСубСост"):
        return seed["contractor_name"], "excel"
    if path.endswith("@НаимРаздел"):
        return seed["section_name"], "excel"
    if path.endswith("@НаимТов"):
        return seed["work_name"], "excel"
    if path.endswith("@ЦенаТов"):
        return seed["price"], "excel"
    if path.endswith("@СтТовБезНДС"):
        return seed["line_total"], "excel"
    if path.endswith("@СтБезНДСРаздОтч"):
        return seed["total"], "excel"
    if path.endswith("@СтТовБезНДСВсего"):
        return seed["total"], "excel"
    if path.endswith("/СумНалВсего"):
        return seed["vat_sum"], "excel"
    if path.endswith("@НалБаза"):
        return seed["total"], "excel"
    if path.endswith("@СодОпер"):
        return seed["operation_text"], "excel"

    # names/organization from excel context by branch
    if path.endswith("@НаимОрг"):
        if "/СвПодр/" in path:
            return seed["contractor_name"], "excel"
        if "/СвЗак/" in path:
            return seed["customer_name"], "excel"
        return seed["builder_name"], "excel"
    if path.endswith("@АдрТекст"):
        if "/МестВыпРаб/" in path:
            return seed["site_name"], "excel"
        if "/СвПодр/" in path:
            return seed["contractor_name"], "excel"
        if "/СвЗак/" in path:
            return seed["customer_name"], "excel"
        return seed["site_name"], "excel"

    # signer names
    if path.endswith("@Фамилия"):
        if "/ПодписантПодр/" in path:
            fam, _, _ = split_name(seed["sign_contractor"])
            return fam, "excel"
        fam, _, _ = split_name(seed["sign_customer"])
        return fam, "excel"
    if path.endswith("@Имя"):
        if "/ПодписантПодр/" in path:
            _, name_part, _ = split_name(seed["sign_contractor"])
            return name_part, "excel"
        _, name_part, _ = split_name(seed["sign_customer"])
        return name_part, "excel"

    # 4) Reference/classifier-like
    if path.endswith("@ЕдИзмПрослеж"):
        return "796", "reference"  # ОКЕИ: штука
    if path.endswith("@КодСтр"):
        return "643", "reference"  # ОКСМ: Россия
    if path.endswith("@КодРегион"):
        return "77", "reference"
    if path.endswith("/КодГАР"):
        return "7700000000000000001", "reference"
    if path.endswith("@ИдСтат"):
        return "ИО", "reference"

    # 5) Enumerations from schema
    if field and field.enum_values:
        return str(field.enum_values[0]), "reference"

    # 6) Heuristics by code/path
    if "ИННЮЛ" in path:
        return "7701234567", "manual"
    if "ИННФЛ" in path:
        return "123456789012", "manual"
    if path.endswith("@Индекс"):
        return "123456", "manual"
    if path.endswith("@НалСт"):
        return "20%", "manual"
    if path.endswith("/СумНал"):
        return seed["vat_sum"], "manual"
    if path.endswith("/БезНДС"):
        return "без НДС", "manual"
    if path.endswith("/ДефНДС"):
        return "-", "manual"
    if path.endswith("/ВидТреб"):
        return "01", "manual"
    if path.endswith("/ВидУдерж"):
        return "31", "manual"
    if path.endswith("@НомТовПрослеж"):
        return "123456789012345678901234567", "manual"
    if path.endswith("@КолВЕдПрослеж"):
        return "1", "manual"
    if path.endswith("/УвелДен") or path.endswith("/УменьшДен"):
        return "1", "manual"
    if path.endswith("/УвелКол") or path.endswith("/УменьшКол"):
        return "1", "manual"
    if path.endswith("/ДаннИно"):
        return "FOREIGN_ORG", "manual"
    if path.endswith("/НаимОИВ"):
        return "Минстрой России", "manual"
    if path.endswith("@Идентиф"):
        return "customField", "manual"
    if path.endswith("@Значение"):
        return "sample", "manual"

    # 7) Pattern-aware fallback
    if field and field.pattern:
        # Very basic generator for common numeric fixed-length patterns
        m = re.fullmatch(r"\\d\{(\d+)\}", field.pattern)
        if m:
            return "1" * int(m.group(1)), "manual"
        if "\\d" in field.pattern:
            return "1", "manual"

    # 8) Generic numeric/text fallback by name
    low = f"{code} {name} {note} {path}".lower()
    if any(k in low for k in ["сум", "стоим", "цена", "кол", "нал", "ндс", "треб", "удерж"]):
        return "1", "manual"

    return "Пример", "manual"


def main() -> int:
    now = datetime.now()

    fields = parse_xsd_fields(XSD_PATH)
    field_by_path = {f.path: f for f in fields}

    mandatory_rows = parse_md_table(MANDATORY_MD, min_cols=6)
    mapped, unmapped = build_path_mapping(mandatory_rows, fields)

    seed = parse_excel_seed(SAMPLE_XLSX)

    values: dict[str, str] = {}
    value_sources: dict[str, str] = {}

    # Fill mapped mandatory leaf fields
    for item in mapped:
        path = item["path"]
        code = item["code"]
        name = item["name"]
        note = item["note"]
        field = field_by_path.get(path)

        value, source = choose_value(path, code, name, note, field, seed, now)

        # Respect max length when present
        if field and field.max_length and value:
            try:
                ml = int(field.max_length)
                value = value[:ml]
            except Exception:
                pass

        # If element has attribute children, we treat it as a container and do not set text.
        if field and field.kind == "element":
            has_attr_children = any(p.startswith(path + "/@") for p in field_by_path.keys())
            if has_attr_children:
                continue

        values[path] = value
        value_sources[path] = source

    # Add a few extra practical fields from sample for better structural completeness
    extras = {
        "/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@НаимДок": "Договор генподряда",
        "/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@НомерДок": seed["contract_no"],
        "/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@ДатаДок": seed["contract_date"],

        "/Файл/Документ/СвАктСдПр/СвПодр/СвСторДог/ИдСв/СвЮЛУч/@НаимОрг": seed["contractor_name"],
        "/Файл/Документ/СвАктСдПр/СвПодр/СвСторДог/ИдСв/СвЮЛУч/@ИННЮЛ": "7701234567",
        "/Файл/Документ/СвАктСдПр/СвЗак/СвСторДог/ИдСв/СвЮЛУч/@НаимОрг": seed["customer_name"],
        "/Файл/Документ/СвАктСдПр/СвЗак/СвСторДог/ИдСв/СвЮЛУч/@ИННЮЛ": "7707654321",

        "/Файл/Документ/НаимИСт/ВидРаб/@НомСтр": "1",
        "/Файл/Документ/НаимИСт/ВидРаб/@НомПоз": "1",
        "/Файл/Документ/НаимИСт/ВидРаб/@ТипЗатр": "1",
        "/Файл/Документ/НаимИСт/ВидРаб/@ОКЕИ_Стройка": "796",
        "/Файл/Документ/НаимИСт/ВидРаб/@НаимЕдИзм": "шт",

        "/Файл/Документ/НаимИСт/Раздел/СвВидРаб/@НаимТов": seed["work_name"],
        "/Файл/Документ/НаимИСт/Раздел/СвВидРаб/@ЦенаТов": seed["price"],
        "/Файл/Документ/НаимИСт/Раздел/СвВидРаб/@СтТовБезНДС": seed["line_total"],

        "/Файл/Документ/ПодписантПодр/Подписант/ФИО/@Фамилия": split_name(seed["sign_contractor"])[0],
        "/Файл/Документ/ПодписантПодр/Подписант/ФИО/@Имя": split_name(seed["sign_contractor"])[1],
        "/Файл/Документ/ВсегоАктОтч/СумПоСтавке/@НалСт": "20%",
        "/Файл/Документ/ВсегоАктОтч/СумПоСтавке/@НалБаза": seed["total"],
        "/Файл/Документ/СвОРасч/@ВсегоКОплатОтч": seed["total"],
    }

    for k, v in extras.items():
        values.setdefault(k, v)
        value_sources.setdefault(k, "excel" if v in seed.values() else "manual")

    # Force some constrained fields (override auto-picked defaults)
    forced = {
        "/Файл/Документ/ВсегоАктОтч/СумПоСтавке/@НалСт": "20%",
        "/Файл/Документ/ВсегоАктОтч/СумПоСтавке/@НалБаза": seed["total"],
        "/Файл/Документ/ВсегоАктОтч/СумПоСтавке/СумНДС": seed["vat_sum"],
        "/Файл/Документ/СвАктСдПр/ИзмСмет/@КодСмет": "1",
        "/Файл/Документ/СвАктСдПр/ИзмСмет/ИдДопСогл/ТипИдДок/@НаимДок": "Дополнительное соглашение",
        "/Файл/Документ/СвАктСдПр/ИзмСмет/ИдДопСогл/ТипИдДок/@НомерДок": "ДС-1",
        "/Файл/Документ/СвАктСдПр/ИзмСмет/ИдДопСогл/ТипИдДок/@ДатаДок": seed["doc_date"],
        "/Файл/Документ/СвАктСдПр/ИспрАктСдПр/@НомИспр": "1",
        "/Файл/Документ/СвАктСдПр/ИспрАктСдПр/@ДатаИспр": seed["doc_date"],
        "/Файл/Документ/СвОРасч/УчетТребУдерж/ВидУдерж": "31",
    }
    for k, v in forced.items():
        values[k] = v
        value_sources[k] = "excel" if v in seed.values() else "manual"

    # Tidy conflicting nodes that frequently break XSD due choice/sequence constraints.
    for conflict_path in [
        # VAT choice inside line item
        "/Файл/Документ/НаимИСт/ВидРаб/СумНал/БезНДС",
        "/Файл/Документ/НаимИСт/ВидРаб/СумНал/ДефНДС",

        # VAT choice inside totals
        "/Файл/Документ/ВсегоАктОтч/ОтсСумНДС",

        # Choice in ИзмененияТип: keep only one quantitative branch
        "/Файл/Документ/НаимИСт/ВидРаб/УчОшИНовОбстСт/ОшибПрПер/УменьшДен",
        "/Файл/Документ/НаимИСт/ВидРаб/УчОшИНовОбстСт/ОшибПрПер/УменьшКол",
        "/Файл/Документ/НаимИСт/ВидРаб/УчОшИНовОбстСт/ОшибПрПер/НетИзмКол",

        # Choice in СумТребУдержТип: keep only one of ВидТреб/ВидУдерж
        "/Файл/Документ/СвОРасч/УчетТребУдерж/ВидТреб",

        # Choice in ИдРекСостТип under ОснДовОргСост
        "/Файл/Документ/ОснДовОргСост/ИдРекСост/ИННФЛ",
        "/Файл/Документ/ОснДовОргСост/ИдРекСост/ДаннИно",
        "/Файл/Документ/ОснДовОргСост/ИдРекСост/ДаннИно/@ИдСтат",
        "/Файл/Документ/ОснДовОргСост/ИдРекСост/НаимОИВ",

        # Choice in participant identification
        "/Файл/Документ/СвАктСдПр/СвЗак/СвСторДог/ИдСв/СвИнНеУч",
        "/Файл/Документ/СвАктСдПр/СвЗак/СвСторДог/ИдСв/СвИнНеУч/@НаимОрг",

        # Choice in address (МестВыпРаб)
        "/Файл/Документ/СвАктСдПр/МестВыпРаб/АдрИнф",
        "/Файл/Документ/СвАктСдПр/МестВыпРаб/АдрИнф/@КодСтр",
        "/Файл/Документ/СвАктСдПр/МестВыпРаб/АдрИнф/@АдрТекст",
        "/Файл/Документ/СвАктСдПр/МестВыпРаб/КодГАР",
    ]:
        values.pop(conflict_path, None)
        value_sources.pop(conflict_path, None)

    root = build_xml_from_values(values)

    OUT_XML.parent.mkdir(parents=True, exist_ok=True)
    OUT_XML.write_bytes(
        __import__("lxml.etree").etree.tostring(
            root,
            encoding="windows-1251",
            pretty_print=True,
            xml_declaration=True,
        )
    )

    errors = validate_xml(root, XSD_PATH)

    # Save full values map for traceability
    OUT_VALUES_JSON.write_text(
        json.dumps(
            {
                "generatedAt": now.isoformat(),
                "mandatoryRows": len(mandatory_rows),
                "mappedLeafRows": len(mapped),
                "unmappedRows": unmapped,
                "values": values,
                "valueSources": value_sources,
                "xsdErrors": errors,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    # Build human report
    src_count = Counter(value_sources.values())
    md: list[str] = []
    md.append("# Пример XML из обязательных данных (1110335) — отчет проверки XSD\n")
    md.append(f"- Сформированный XML: `{OUT_XML.relative_to(ROOT)}`")
    md.append(f"- Источник обязательных полей: `{MANDATORY_MD.relative_to(ROOT)}`")
    md.append(f"- Источник бизнес-данных: `{SAMPLE_XLSX.relative_to(ROOT)}`")
    md.append(f"- Дата/время генерации: `{now.strftime('%Y-%m-%d %H:%M:%S')}`\n")

    md.append("## Что сделано")
    md.append(f"- Обязательных строк в списке: **{len(mandatory_rows)}**")
    md.append(f"- Удалось сопоставить с XML-путями (leaf): **{len(mapped)}**")
    md.append(f"- Не сопоставилось напрямую (в основном групповые узлы): **{len(unmapped)}**\n")

    md.append("## Источники значений")
    md.append(f"- Из Excel: **{src_count.get('excel', 0)}**")
    md.append(f"- Сгенерировано автоматически: **{src_count.get('generated', 0)}**")
    md.append(f"- Справочники/классификаторы: **{src_count.get('reference', 0)}**")
    md.append(f"- Ручная генерация (по формату): **{src_count.get('manual', 0)}**\n")

    md.append("## Результат XSD-валидации")
    md.append(f"- Ошибок: **{len(errors)}**")
    md.append("- Статус: **FAILED**" if errors else "- Статус: **PASS**")

    if errors:
        md.append("\n### Список несоответствий / ошибок")
        for i, e in enumerate(errors, 1):
            md.append(f"{i}. `{e}`")

    if unmapped:
        md.append("\n### Поля из обязательного списка, которые не удалось напрямую сопоставить с leaf-путем")
        md.append("(обычно это групповые/контейнерные элементы)")
        md.append("| Таблица | Код | Поле | Причина |")
        md.append("|---|---|---|---|")
        for u in unmapped:
            tab = str(u.get("tab", "")).replace("|", "\\|")
            code = str(u.get("code", "")).replace("|", "\\|")
            name = str(u.get("name", "")).replace("|", "\\|")
            reason = str(u.get("reason", "")).replace("|", "\\|")
            md.append(f"| {tab} | {code or '—'} | {name} | {reason} |")

    OUT_REPORT.write_text("\n".join(md), encoding="utf-8")

    print(f"XML written: {OUT_XML}")
    print(f"Report written: {OUT_REPORT}")
    print(f"XSD errors: {len(errors)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
