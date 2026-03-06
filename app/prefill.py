from __future__ import annotations

from pathlib import Path
from typing import Dict
import datetime as dt
import openpyxl


def _fmt_date(v) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, dt.date):
        return v.strftime("%d.%m.%Y")
    return str(v).strip()


def prefill_from_first_sheet(xlsx_path: Path) -> Dict[str, str]:
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.worksheets[0]

    data: Dict[str, str] = {}

    # Header-like fields from your sample layout
    now = dt.datetime.now()
    data["/Файл/@ВерсФорм"] = "1.00"
    data["/Файл/@ВерсПрог"] = "KC2-KC3-XML-webapp"
    data["/Файл/Документ/@КНД"] = "1115003"
    data["/Файл/Документ/@ДатаИнфПодр"] = now.strftime("%d.%m.%Y")
    data["/Файл/Документ/@ВремИнфПодр"] = now.strftime("%H.%M.%S")

    doc_num = ws.cell(23, 7).value
    doc_date = ws.cell(23, 8).value
    obj = ws.cell(15, 4).value
    contractor = ws.cell(11, 4).value

    if doc_num is not None:
        data["/Файл/Документ/СвАктСдПр/@НомерДок"] = str(doc_num).strip()
    if doc_date is not None:
        data["/Файл/Документ/СвАктСдПр/@ДатаДок"] = _fmt_date(doc_date)
    if obj is not None:
        data["/Файл/Документ/СвАктСдПр/@НаимОб"] = str(obj).strip()
    if contractor is not None:
        data["/Файл/Документ/@НаимЭкСубСост"] = str(contractor).strip()

    # Contract number/date (row 17/18 in sample)
    contract_no = ws.cell(17, 10).value
    contract_dt = ws.cell(18, 10).value
    if contract_no:
        data["/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@НомерДок"] = str(contract_no).strip()
    if contract_dt:
        data["/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@ДатаДок"] = _fmt_date(contract_dt)
    data["/Файл/Документ/СвАктСдПр/ИдДог/ТипИдДок/@НаимДок"] = "Договор генподряда"

    # Basis text
    basis = ws.cell(25, 3).value
    if basis:
        data["/Файл/Документ/НаимИСт"] = str(basis).strip()

    return data
